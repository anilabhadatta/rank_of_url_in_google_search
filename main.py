from flask import Flask, render_template, Response, request, redirect, url_for
import requests
from bs4 import BeautifulSoup
from werkzeug.utils import secure_filename
import os
from openpyxl import load_workbook
from fake_useragent import UserAgent
import random

app = Flask(__name__)


@app.route("/")
def index():
    return render_template("index.html", condition=False)


@app.route("/result/", methods=['POST'])
def search():
    excel_file = request.files['file']
    file_name = secure_filename(excel_file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)
    excel_file.save(file_path)
    keyword_list, webpage_list, rank_list = [], [], []
    read_excel(file_path, file_name, keyword_list, webpage_list, rank_list)
    return render_template("index.html", len=len(keyword_list), keyword_list=keyword_list, webpage_list=webpage_list, rank_list=rank_list, condition=True)


def read_excel(file_path, file_name, keyword_list, webpage_list, rank_list):
    print("Reading Excel")
    excel_path = file_path
    work_book = load_workbook(excel_path, read_only=False)
    work_sheet = work_book["Sheet1"]
    print(work_sheet.max_row+1)
    for row in range(2, work_sheet.max_row+1):
        keyword = work_sheet.cell(row=row, column=1)
        check_url = work_sheet.cell(row=row, column=2)
        rank_cell = work_sheet.cell(row=row, column=3)

        print("Searching :", keyword.value, check_url.value)
        if keyword.value == None or check_url.value == None:
            print("Skipping , invalid parameters")
            rank = "-1,"
            rank_cell.value = rank[:-1]
            pass
        else:
            rank = find_rank(keyword.value, check_url.value)
            rank_cell.value = rank[:-1]

        keyword_list.append(keyword.value)
        webpage_list.append(check_url.value)
        rank_list.append(rank[:-1])

    newfile = os.getcwd()+"\\"+file_name
    work_book.save(newfile)
    os.remove(file_path)


def find_rank(keyword, check_url):
    #proxies = ['14.140.131.82:3128']
    #proxy = random.choice(proxies)
    #print(proxy)
    number_result, rank = 100, 1
    ua = UserAgent()
    url = "https://www.google.com/search?q="
    keyword = keyword.replace(" ", "+")
    url = url + keyword + "&num=" + str(number_result)
    #page = requests.get(url,proxies={"http": proxy, "https": proxy},headers={"User-Agent": ua.random})
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'html.parser')
    result_div = soup.find_all(
        'div', attrs={'class': 'ZINbbc xpd O9g5cc uUPGi'})
    rank_list_temp = ""

    # debugging html file
    # f = open("output.html",'w',encoding="utf-8")
    # f.write(page.text)
    # f.close()

    for div in result_div:
        try:
            link = div.find("a", href=True)
            # print(link['href'][7:7+len(check_url)], rank)
            if link['href'][7:7+len(check_url)] == check_url:
                print("Found Website , Rank :", rank)
                rank_list_temp += str(rank)+","
            rank += 1
        except:
            pass
    return (rank_list_temp,"Website not found")[rank_list_temp == ""]


if __name__ == "__main__":
    home_dir = os.path.expanduser("~")
    UPLOAD_FOLDER = os.path.join(home_dir, "upload\\")
    app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
    try:
        os.makedirs(UPLOAD_FOLDER)
    except:
        pass
    app.run(debug=True)
